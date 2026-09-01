#!/usr/bin/env python3
"""Convert the user-supplied dealer workbook into a validated JSON directory."""

from __future__ import annotations

import argparse
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


EXPECTED_HEADERS = ["DEALER", "DEALER CODE", "OPERATIONAL AREA"]


def clean(value: object) -> str:
    return "" if value is None else str(value).strip()


def clean_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("input", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook_bytes = args.input.read_bytes()
    workbook = load_workbook(args.input, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A planilha oficial de dealers está vazia.")

    headers = [clean(value).upper() for value in rows[0][:3]]
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Cabeçalhos inválidos: {headers}")

    dealers = []
    seen_names: set[str] = set()
    seen_codes: set[str] = set()
    for row_number, row in enumerate(rows[1:], start=2):
        name = clean(row[0] if len(row) > 0 else None)
        code = clean_code(row[1] if len(row) > 1 else None)
        operational_area = clean(row[2] if len(row) > 2 else None)
        if not name and not code and not operational_area:
            continue
        if not name:
            raise ValueError(f"Dealer sem nome na linha {row_number}.")
        name_key = name.casefold()
        if name_key in seen_names:
            raise ValueError(f"Dealer duplicado: {name}")
        if code and code in seen_codes:
            raise ValueError(f"Código de dealer duplicado: {code}")
        seen_names.add(name_key)
        if code:
            seen_codes.add(code)
        dealers.append(
            {
                "name": name,
                "code": code or None,
                "operationalArea": operational_area or None,
            }
        )

    payload = {
        "sourceWorkbook": args.input.name,
        "sourceSheet": sheet.title,
        "sourceHeaders": EXPECTED_HEADERS,
        "sourceSha256": hashlib.sha256(workbook_bytes).hexdigest(),
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "dealerCount": len(dealers),
        "dealers": dealers,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
