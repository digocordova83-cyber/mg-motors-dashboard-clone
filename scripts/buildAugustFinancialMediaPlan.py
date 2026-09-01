from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = Path('/home/ubuntu/mg-motors-dashboard-clone/exports/MGBrazil_DigitalMediaPlan_August2026_Financial_Structure.xlsx')
SOURCE = Path('/home/ubuntu/upload/pasted_file_6fhRVV_MGBrazil-DigitalMediaPlanAug26-MediaSAVE99k-Harmonized-EN.xlsx')
OUT.parent.mkdir(parents=True, exist_ok=True)

# August values are taken from the uploaded August media plan. Actual Investment is
# intentionally left blank because no August payment/proof data was supplied.
commission_rate = 0.04
line_up = [
    ('Digital', 'Google Ads', 'Google', 350000, 'August media plan • Consolidated Google'),
    ('Digital', 'Webmotors', 'Webmotors', 105000, 'August media plan • R$ 120k total including Urban'),
    ('Digital', 'Publya Programmatic Display', 'Publya', 55000, 'August media plan • Geolocation'),
    ('Digital', 'Meta Ads', 'Publya', 150000, 'August media plan • Line-up allocation'),
    ('Digital', 'TikTok Ads — Test', 'TikTok', 30000, 'August media plan • Test allocation'),
    ('Digital', 'Mercado Livre Ads', 'Mercado Livre', 61000, 'August media plan • Marketplace conversion'),
]
urban = [
    ('Digital', 'YouTube Video', 'Publya', 35000, 'August media plan • Upper funnel / launch'),
    ('Digital', 'Google Ads', 'Google', 80000, 'August media plan • MG4 Urban'),
    ('Digital', 'Meta Ads', 'Publya', 60000, 'August media plan • MG4 Urban'),
    ('Digital', 'Webmotors', 'Webmotors', 15000, 'August media plan • MG4 Urban'),
    ('Digital', 'Mercado Livre Ads', 'Mercado Livre', 10000, 'August media plan • MG4 Urban'),
]
magazine = [
    ('OFF', 'Magazine', 'TBD', 50000, 'August media plan • Outlet / activation not yet defined'),
]
media_save = 99000

wb = Workbook()
ws = wb.active
ws.title = 'Financial Plan - Aug26'
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'B10'
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.orientation = 'landscape'
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins.left = 0.2
ws.page_margins.right = 0.2
ws.page_margins.top = 0.35
ws.page_margins.bottom = 0.35
ws.print_title_rows = '1:10'

NAVY = '09283D'
NAVY_2 = '123B55'
RED = 'D7194F'
GOLD = 'D6A927'
TEAL = '19A7A7'
LIGHT_BLUE = 'E8F0F6'
LIGHT_GREEN = 'E6F4EE'
LIGHT_YELLOW = 'FFF2CC'
LIGHT_RED = 'FCE4EC'
GREY = 'F3F5F7'
MID_GREY = 'D9E0E6'
WHITE = 'FFFFFF'
BLACK = '1F2933'
GREEN = '0B6E4F'
ORANGE = 'B45F06'

thin_grey = Side(style='thin', color='D7DEE5')
medium_navy = Side(style='medium', color=NAVY)
section_border = Border(top=medium_navy, bottom=medium_navy)
cell_border = Border(left=thin_grey, right=thin_grey, top=thin_grey, bottom=thin_grey)

for col, width in {'A':3, 'B':14, 'C':28, 'D':18, 'E':18, 'F':19, 'G':18, 'H':20, 'I':18, 'J':48}.items():
    ws.column_dimensions[col].width = width

for row in range(1, 80):
    ws.row_dimensions[row].height = 20
ws.row_dimensions[2].height = 32
ws.row_dimensions[3].height = 24

# Main heading
ws.merge_cells('B2:J2')
ws['B2'] = 'MG MOTOR BRAZIL — FINANCIAL MEDIA PLAN BY CHANNEL | AUGUST 2026'
ws['B2'].font = Font(name='Georgia', size=16, bold=True, color=WHITE)
ws['B2'].fill = PatternFill('solid', fgColor=NAVY)
ws['B2'].alignment = Alignment(horizontal='center', vertical='center')

ws.merge_cells('B3:J3')
ws['B3'] = 'Line-up BRL 751,000 • Urban BRL 200,000 • Off / Magazine BRL 50,000 • Media Save BRL 99,000 • Consolidated gross BRL 1,100,000'
ws['B3'].font = Font(name='Calibri', size=10, bold=True, color=WHITE)
ws['B3'].fill = PatternFill('solid', fgColor=RED)
ws['B3'].alignment = Alignment(horizontal='center', vertical='center')

# Assumption row
ws['B5'] = 'COMMISSION RATE'
ws['B5'].font = Font(name='Calibri', size=10, bold=True, color=WHITE)
ws['B5'].fill = PatternFill('solid', fgColor=NAVY)
ws['B5'].alignment = Alignment(horizontal='center', vertical='center')
ws['C5'] = commission_rate
ws['C5'].number_format = '0.0%'
ws['C5'].font = Font(name='Calibri', size=12, bold=True, color=NAVY)
ws['C5'].fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
ws['C5'].alignment = Alignment(horizontal='center')
ws.merge_cells('D5:J5')
ws['D5'] = 'Gross = planned media value | Commission = Gross × rate | Net = Gross − Commission | Actual Investment intentionally blank until August payment evidence is available'
ws['D5'].font = Font(name='Calibri', size=9, italic=True, color='5B6570')
ws['D5'].alignment = Alignment(wrap_text=True, vertical='center')

# KPI strip references consolidated total row 15
kpis = [
    ('B7', 'TOTAL GROSS', 'E15', 'R$ #,##0.00'),
    ('D7', 'TOTAL COMMISSION', 'F15', 'R$ #,##0.00'),
    ('F7', 'TOTAL NET', 'G15', 'R$ #,##0.00'),
    ('H7', 'TOTAL BUDGET STATUS', 'I15', '@'),
]
for label_cell, label, value_ref, fmt in kpis:
    label_col = label_cell[0]
    value_col = get_column_letter(ord(label_col) - 64 + 1)
    ws[label_cell] = label
    ws[label_cell].font = Font(name='Calibri', size=9, bold=True, color=WHITE)
    ws[label_cell].fill = PatternFill('solid', fgColor=GOLD if label != 'TOTAL BUDGET STATUS' else RED)
    ws[label_cell].alignment = Alignment(horizontal='center', vertical='center')
    ws[value_col + '7'] = f'={value_ref}'
    ws[value_col + '7'].font = Font(name='Calibri', size=11, bold=True, color=NAVY)
    ws[value_col + '7'].fill = PatternFill('solid', fgColor=LIGHT_BLUE)
    ws[value_col + '7'].number_format = fmt
    ws[value_col + '7'].alignment = Alignment(horizontal='center', vertical='center')
# status KPI overrides because it should be textual
ws['I7'] = '=I15'
ws['I7'].number_format = '@'

headers = ['CATEGORY', 'CHANNEL', 'PUBLISHER', 'PLAN GROSS (BRL)', '4% COMMISSION (BRL)', 'NET (BRL)', 'ACTUAL INVESTMENT (BRL)', 'STATUS', 'SOURCE / NOTE']

def style_header(row):
    for col in range(2, 11):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.font = Font(name='Calibri', size=9, bold=True, color=WHITE)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = cell_border
    ws.row_dimensions[row].height = 30

def style_section(row, title, fill=RED):
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    cell = ws.cell(row=row, column=2)
    cell.value = title
    cell.fill = PatternFill('solid', fgColor=fill)
    cell.font = Font(name='Georgia', size=12, bold=True, color=WHITE)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = section_border
    ws.row_dimensions[row].height = 24

def write_table_header(row):
    for col, header in enumerate(headers, start=2):
        ws.cell(row=row, column=col, value=header)
    style_header(row)

def write_detail_rows(start_row, rows, source_type='plan'):
    for idx, (category, channel, publisher, gross, note) in enumerate(rows, start=start_row):
        ws.cell(idx, 2, category)
        ws.cell(idx, 3, channel)
        ws.cell(idx, 4, publisher)
        ws.cell(idx, 5, gross)
        ws.cell(idx, 6, f'=E{idx}*$C$5')
        ws.cell(idx, 7, f'=E{idx}-F{idx}')
        ws.cell(idx, 8, None)  # Actual investment: no August payment evidence provided.
        ws.cell(idx, 9, 'PENDING')
        ws.cell(idx, 10, note)
        for col in range(2, 11):
            cell = ws.cell(idx, col)
            cell.border = cell_border
            cell.alignment = Alignment(vertical='center', wrap_text=(col in (3, 4, 10)))
            cell.font = Font(name='Calibri', size=9, color=BLACK)
            if idx % 2 == 0:
                cell.fill = PatternFill('solid', fgColor=GREY)
        ws.cell(idx, 8).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
        ws.cell(idx, 8).number_format = 'R$ #,##0.00'
        ws.cell(idx, 9).fill = PatternFill('solid', fgColor=LIGHT_RED)
        ws.cell(idx, 9).font = Font(name='Calibri', size=9, bold=True, color=ORANGE)
        for col in (5, 6, 7):
            ws.cell(idx, col).number_format = 'R$ #,##0.00'
    return start_row + len(rows) - 1

def subtotal(row, label, first_data, last_data, note):
    ws.cell(row, 2, 'SUBTOTAL')
    ws.cell(row, 3, label)
    ws.cell(row, 5, f'=SUM(E{first_data}:E{last_data})')
    ws.cell(row, 6, f'=SUM(F{first_data}:F{last_data})')
    ws.cell(row, 7, f'=SUM(G{first_data}:G{last_data})')
    ws.cell(row, 8, None)
    ws.cell(row, 9, 'PENDING')
    ws.cell(row, 10, note)
    for col in range(2, 11):
        cell = ws.cell(row, col)
        cell.fill = PatternFill('solid', fgColor=NAVY)
        cell.font = Font(name='Calibri', size=9, bold=True, color=WHITE)
        cell.border = cell_border
        cell.alignment = Alignment(vertical='center', wrap_text=(col in (3, 10)))
    for col in (5, 6, 7):
        ws.cell(row, col).number_format = 'R$ #,##0.00'
    ws.cell(row, 8).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    ws.cell(row, 8).number_format = 'R$ #,##0.00'
    ws.cell(row, 9).fill = PatternFill('solid', fgColor=LIGHT_RED)
    return row

# Consolidated summary block
style_section(9, 'CONSOLIDATED SUMMARY | AUGUST 2026', NAVY_2)
write_table_header(10)
summary_rows = [
    ('DIGITAL', 'Digital Line-up', 'See detail below', '=E25', '=F25', '=G25', None, 'PLANNED', 'Consolidated from Digital Line-up subtotal'),
    ('DIGITAL', 'Digital Urban', 'See detail below', '=E34', '=F34', '=G34', None, 'PLANNED', 'Dedicated MG4 Urban budget'),
    ('DIGITAL', 'MEDIA SAVE', 'Tactical reserve', media_save, '=0', '=E13-F13', None, 'AVAILABLE', 'Reserved within August digital plan; not direct channel allocation'),
    ('OFF', 'Magazine', 'TBD', '=E39', '=F39', '=G39', None, 'PLANNED', 'Off / Magazine block below'),
]
for row, (category, channel, publisher, gross, comm, net, actual, status, note) in enumerate(summary_rows, start=11):
    vals = [category, channel, publisher, gross, comm, net, actual, status, note]
    for col, val in enumerate(vals, start=2):
        ws.cell(row, col, val)
    for col in range(2, 11):
        cell = ws.cell(row, col)
        cell.border = cell_border
        cell.font = Font(name='Calibri', size=9, color=BLACK)
        cell.alignment = Alignment(vertical='center', wrap_text=(col in (3, 4, 10)))
        if row % 2 == 1:
            cell.fill = PatternFill('solid', fgColor=GREY)
    for col in (5, 6, 7):
        ws.cell(row, col).number_format = 'R$ #,##0.00'
    ws.cell(row, 8).fill = PatternFill('solid', fgColor=LIGHT_YELLOW)
    ws.cell(row, 8).number_format = 'R$ #,##0.00'
    ws.cell(row, 9).fill = PatternFill('solid', fgColor=LIGHT_GREEN if status == 'AVAILABLE' else LIGHT_RED)
    ws.cell(row, 9).font = Font(name='Calibri', size=9, bold=True, color=GREEN if status == 'AVAILABLE' else ORANGE)
subtotal(15, 'TOTAL CONSOLIDATED', 11, 14, 'Allocated digital + Urban + Media Save + Off / Magazine')
ws['H15'] = 'PENDING'
ws['I15'] = 'No August actual payment/proof data was provided; blank Actual Investment cells are not zero.'

# Line-up block
style_section(17, 'DIGITAL LINE-UP | Planned gross: BRL 751,000', RED)
write_table_header(18)
line_last = write_detail_rows(19, line_up)
subtotal(25, 'DIGITAL LINE-UP', 19, line_last, 'Performance media allocation from August workbook')

# Urban block
style_section(27, 'DIGITAL URBAN | Planned gross: BRL 200,000', RED)
write_table_header(28)
urban_last = write_detail_rows(29, urban)
subtotal(34, 'DIGITAL URBAN', 29, urban_last, 'Dedicated MG4 Urban allocation from August workbook')

# Magazine block
style_section(36, 'OFF — MAGAZINE | Planned gross: BRL 50,000', RED)
write_table_header(37)
mag_last = write_detail_rows(38, magazine)
subtotal(39, 'OFF — MAGAZINE', 38, mag_last, 'Off / Magazine reserve; outlet and activation remain TBD')

# Notes and audit trail
ws.merge_cells('B42:J42')
ws['B42'] = 'READING GUIDE'
ws['B42'].fill = PatternFill('solid', fgColor=NAVY)
ws['B42'].font = Font(name='Georgia', size=11, bold=True, color=WHITE)
ws['B42'].alignment = Alignment(horizontal='left')
notes = [
    'Source: uploaded MGBrazil-DigitalMediaPlanAug26-MediaSAVE99k-Harmonized-EN.xlsx, August 2026 version.',
    'Digital Line-up = BRL 751,000; Digital Urban = BRL 200,000; Media Save = BRL 99,000; Off / Magazine = BRL 50,000.',
    'Consolidated gross = BRL 1,100,000; commission is formula-driven at 4%; consolidated net = BRL 1,059,960.',
    'Actual Investment cells are intentionally blank because no August payment/proof values were supplied. Blank means not provided, not zero.',
    'Projected lead target in the August source workbook: 12,000 Leads. This financial workbook does not alter operational dashboard volumes.',
]
for i, note in enumerate(notes, start=43):
    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=10)
    ws.cell(i, 2, '• ' + note)
    ws.cell(i, 2).font = Font(name='Calibri', size=9, italic=True, color='5B6570')
    ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='center')
    ws.row_dimensions[i].height = 22

# Status validation and conditional formatting
status_dv = DataValidation(type='list', formula1='"PLANNED,PENDING,PAID,AVAILABLE,PARTIALLY PAID,TBD"', allow_blank=True)
ws.add_data_validation(status_dv)
status_dv.add('I11:I39')
ws.conditional_formatting.add('I11:I39', FormulaRule(formula=['I11="PAID"'], fill=PatternFill('solid', fgColor=LIGHT_GREEN)))
ws.conditional_formatting.add('I11:I39', FormulaRule(formula=['I11="PENDING"'], fill=PatternFill('solid', fgColor=LIGHT_RED)))

# Borders around key rows and alignment
for row in (9, 17, 27, 36, 42):
    for col in range(2, 11):
        ws.cell(row, col).border = section_border
for row in (15, 25, 34, 39):
    ws.row_dimensions[row].height = 23
for row in range(1, 50):
    for col in range(2, 11):
        ws.cell(row, col).alignment = ws.cell(row, col).alignment.copy(vertical='center')

# Documentation sheet with source and definitions
readme = wb.create_sheet('Source & Assumptions')
readme.sheet_view.showGridLines = False
readme.column_dimensions['A'].width = 3
readme.column_dimensions['B'].width = 28
readme.column_dimensions['C'].width = 85
readme.merge_cells('B2:C2')
readme['B2'] = 'AUGUST 2026 | SOURCE & ASSUMPTIONS'
readme['B2'].fill = PatternFill('solid', fgColor=NAVY)
readme['B2'].font = Font(name='Georgia', size=15, bold=True, color=WHITE)
readme['B2'].alignment = Alignment(horizontal='center')
source_rows = [
    ('Source workbook', SOURCE.name),
    ('Period', 'August 2026'),
    ('Commission rate', '4.0%, consistent with the July reference structure'),
    ('Digital Line-up gross', 'BRL 751,000'),
    ('Digital Urban gross', 'BRL 200,000'),
    ('Media Save reserve', 'BRL 99,000; no commission applied in the source plan'),
    ('Off / Magazine gross', 'BRL 50,000; outlet and activation not yet defined'),
    ('Consolidated gross', 'BRL 1,100,000'),
    ('Consolidated commission', 'BRL 40,040 at 4%, excluding Media Save'),
    ('Consolidated net', 'BRL 1,059,960'),
    ('Projected Leads target', '12,000 in the August source workbook'),
    ('Actual Investment', 'Not provided for August in the supplied workbook; cells intentionally remain blank and status is Pending.'),
]
for r, (k, v) in enumerate(source_rows, start=5):
    readme.cell(r, 2, k)
    readme.cell(r, 3, v)
    readme.cell(r, 2).font = Font(name='Calibri', size=10, bold=True, color=NAVY)
    readme.cell(r, 3).font = Font(name='Calibri', size=10, color=BLACK)
    readme.cell(r, 2).fill = PatternFill('solid', fgColor=LIGHT_BLUE if r % 2 else GREY)
    readme.cell(r, 3).fill = PatternFill('solid', fgColor=LIGHT_BLUE if r % 2 else GREY)
    readme.cell(r, 2).border = cell_border
    readme.cell(r, 3).border = cell_border
    readme.cell(r, 3).alignment = Alignment(wrap_text=True, vertical='center')
    readme.row_dimensions[r].height = 23
readme.merge_cells('B20:C20')
readme['B20'] = 'No numerical value was fabricated for August Actual Investment. Update the yellow cells on the main sheet when payment/proof values are available.'
readme['B20'].font = Font(name='Calibri', size=10, italic=True, color=ORANGE)
readme['B20'].alignment = Alignment(wrap_text=True, vertical='center')
readme.row_dimensions[20].height = 34

# Workbook metadata
wb.properties.title = 'MG Brazil — August 2026 Financial Media Plan'
wb.properties.subject = 'Digital Line-up, Digital Urban, Off Magazine and Consolidated Summary'
wb.properties.creator = 'MG Motors Dashboard'
wb.properties.description = 'August 2026 financial media structure based on the supplied August media plan workbook.'
wb.save(OUT)
print(OUT)
