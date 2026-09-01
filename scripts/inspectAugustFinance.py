from pathlib import Path
from openpyxl import load_workbook

path = Path('/home/ubuntu/upload/pasted_file_6fhRVV_MGBrazil-DigitalMediaPlanAug26-MediaSAVE99k-Harmonized-EN.xlsx')
for data_only in (False, True):
    wb = load_workbook(path, data_only=data_only)
    print('DATA_ONLY', data_only)
    for sheet_name in ('Resume Plan', 'Media Plan - Digital', 'Media SAVE', 'Budget Split'):
        ws = wb[sheet_name]
        print(f'SHEET {sheet_name}')
        for r in range(1, ws.max_row + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 16) + 1)]
            if any(v not in (None, '') for v in values):
                print(f'{r}: ' + ' || '.join('' if v is None else str(v) for v in values))
    print()
