from __future__ import annotations

import argparse
import json
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


def decimal_value(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal(0)
    if isinstance(value, str):
        normalized = value.replace("R$", "").replace(" ", "").replace(".", "").replace(",", ".")
        return Decimal(normalized)
    return Decimal(str(value))


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    args = parser.parse_args()

    workbook_values = load_workbook(args.workbook, data_only=True, read_only=False)
    workbook_formulas = load_workbook(args.workbook, data_only=False, read_only=False)
    sheet = workbook_values[workbook_values.sheetnames[0]]
    formula_sheet = workbook_formulas[workbook_formulas.sheetnames[0]]

    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    formula_cells: list[str] = []
    channels = ["GOOGLE", "META", "PUBLYA", "WEBMOTORS", "MERCADO LIVRE", "TIKTOK"]

    for row_number in range(2, sheet.max_row + 1):
        values = {headers[index]: sheet.cell(row_number, index + 1).value for index in range(len(headers))}
        if not str(values.get("DEALER") or "").strip():
            continue
        formulas = {
            headers[index]: formula_sheet.cell(row_number, index + 1).value
            for index in range(len(headers))
            if formula_sheet.cell(row_number, index + 1).data_type == "f"
        }
        formula_cells.extend(f"{formula_sheet.cell(row_number, headers.index(header) + 1).coordinate}:{formula}" for header, formula in formulas.items())
        channel_total = sum(decimal_value(values.get(channel)) for channel in channels)
        total_dealer = decimal_value(values.get("TOTAL DEALER"))
        rows.append({
            "row": row_number,
            "dealer": str(values["DEALER"]).strip(),
            "channels": {channel: int(decimal_value(values.get(channel))) for channel in channels},
            "channelTotal": int(channel_total),
            "totalDealer": int(total_dealer),
            "sales": int(decimal_value(values.get("SALES"))),
            "weightPercent": float(decimal_value(values.get("WEIGHT")) * 100),
            "conversionInvestment": float(decimal_value(values.get("CONVERSION INVESTMENT"))),
            "reconcilesChannels": channel_total == total_dealer,
            "formulas": formulas,
        })

    total_dealer = sum(int(row["totalDealer"]) for row in rows)
    total_sales = sum(int(row["sales"]) for row in rows)
    total_weight = sum(Decimal(str(row["weightPercent"])) for row in rows)
    channel_totals = {
        channel: sum(int(row["channels"][channel]) for row in rows)  # type: ignore[index]
        for channel in channels
    }
    report = {
        "workbook": str(args.workbook),
        "sheet": sheet.title,
        "dimensions": {"rows": sheet.max_row, "columns": sheet.max_column},
        "headers": headers,
        "dealerRows": len(rows),
        "hasExplicitCompetence": any("COMPET" in header.upper() or "MÊS" in header.upper() or "MES" in header.upper() for header in headers),
        "formulaCellCount": len(formula_cells),
        "formulaCells": formula_cells,
        "totals": {
            "totalDealer": total_dealer,
            "sales": total_sales,
            "weightPercent": float(total_weight),
            "conversionInvestment": round(sum(float(row["conversionInvestment"]) for row in rows), 2),
            "channels": channel_totals,
            "channelsGrandTotal": sum(channel_totals.values()),
        },
        "checks": {
            "allRowsReconcileChannels": all(bool(row["reconcilesChannels"]) for row in rows),
            "grandTotalReconcilesChannels": sum(channel_totals.values()) == total_dealer,
            "weightsApprox100": abs(float(total_weight) - 100) <= 0.1,
        },
        "rows": rows,
    }
    args.output.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps(report["totals"], ensure_ascii=False))


if __name__ == "__main__":
    main()
