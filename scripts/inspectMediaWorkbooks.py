from pathlib import Path
from openpyxl import load_workbook

paths = [
    Path('/home/ubuntu/upload/pasted_file_BjRAZr_MGBrazil-DigitalMediaPlanJul26(EN)1.xlsx'),
    Path('/home/ubuntu/upload/pasted_file_lhnEJS_MGBrazil-PlanejadoDigital-Jul2026-Gross-Comissao-Liquido-InvestidoReal(2).xlsx'),
    Path('/home/ubuntu/upload/pasted_file_6fhRVV_MGBrazil-DigitalMediaPlanAug26-MediaSAVE99k-Harmonized-EN.xlsx'),
]
for path in paths:
    print(f'FILE: {path.name}')
    if not path.exists():
        print('MISSING')
        continue
    wb = load_workbook(path, data_only=False, read_only=False)
    print('SHEETS:', ' | '.join(wb.sheetnames))
    for ws in wb.worksheets:
        nonempty = []
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), min_col=1, max_col=min(ws.max_column, 16)):
            vals = [cell.value for cell in row]
            if any(v not in (None, '') for v in vals):
                nonempty.append((row[0].row, vals))
        print(f'  SHEET {ws.title}: rows={ws.max_row} cols={ws.max_column} nonempty_preview={len(nonempty)}')
        for row_num, vals in nonempty[:12]:
            compact = [str(v)[:90] if v is not None else '' for v in vals]
            print(f'    {row_num}: ' + ' || '.join(compact))
    print()
